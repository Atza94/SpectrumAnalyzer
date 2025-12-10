#!/usr/bin/env python3
"""
Dual Spectrum Analyzer Viewer - Professional GUI
Connects to two DSA1030 Spectrum Analyzers via VISA/SCPI
"""

import time
import threading
import customtkinter as ctk
from tkinter import messagebox, filedialog
import csv
from datetime import datetime
import webbrowser
import os

import numpy as np
import pyvisa

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DEFAULT_IP1 = "192.168.60.70"
DEFAULT_IP2 = "192.168.60.69"

YMIN_DBM = -120
YMAX_DBM = 0

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class HelpDialog(ctk.CTkToplevel):
    """Help dialog with application information"""
    def __init__(self, parent):
        super().__init__(parent)
        
        self.title("Help & Information")
        self.geometry("550x580")
        self.resizable(False, False)
        
        self.configure(fg_color="#0a0a1a")
        
        self.transient(parent)
        self.after(100, self._set_focus)
        
        header = ctk.CTkFrame(self, fg_color="#16213e", corner_radius=0, height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text="Dual Spectrum Analyzer Viewer",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#00d4ff"
        ).pack(side="left", padx=20, pady=10)
        
        ctk.CTkLabel(
            header,
            text="v1.0",
            font=ctk.CTkFont(size=12),
            text_color="#666688"
        ).pack(side="right", padx=20, pady=10)
        
        content = ctk.CTkScrollableFrame(self, fg_color="#0a0a1a")
        content.pack(fill="both", expand=True, padx=20, pady=20)
        
        self._add_section(content, "About", 
            "This application connects to two Rigol DSA1030 Spectrum Analyzers "
            "via VISA/SCPI protocol over TCP/IP. It provides real-time spectrum "
            "visualization with peak hold functionality."
        )
        
        self._add_section(content, "Getting Started",
            "1. Enter the IP address of your spectrum analyzer\n"
            "2. Click 'Connect' to establish connection\n"
            "3. Set the Center Frequency and Span (in MHz)\n"
            "4. Select the number of sweep Points\n"
            "5. Click 'Single Sweep' or 'Continuous' to start"
        )
        
        self._add_section(content, "Controls",
            "Connect/Disconnect - Toggle connection to the analyzer\n"
            "Single Sweep - Perform one sweep and display results\n"
            "Continuous - Continuously sweep and update the display\n"
            "Reset Peak - Clear the peak hold (yellow) trace\n"
            "Points - Number of data points per sweep (fewer = faster)"
        )
        
        self._add_section(content, "Interactive Marker",
            "Click anywhere on the spectrum plot to place a marker. "
            "Drag to move it along the spectrum. The marker shows:\n\n"
            "- Frequency (MHz) at the marker position\n"
            "- Current amplitude (dBm)\n"
            "- Max Hold amplitude (dBm) if peak tracking is active\n\n"
            "Marker info is displayed in the bottom panel for quick reference."
        )
        
        self._add_section(content, "Collapsible Panels",
            "Click the ▼/▶ arrow button on each control panel to collapse or "
            "expand it. Collapsing a panel hides the connection and sweep settings, "
            "allowing more space for the spectrum plots."
        )
        
        self._add_section(content, "Spectrum Display",
            "Green line - Current sweep measurement\n"
            "Yellow line - Peak hold (maximum values across sweeps)\n"
            "Magenta marker - Interactive frequency/amplitude cursor\n\n"
            "Bottom panel shows:\n"
            "• SA1 marker info (green) - Frequency, amplitude, and max hold\n"
            "• SA2 marker info (pink) - Same data for the second analyzer"
        )
        
        self._add_section(content, "Network Requirements",
            "Your computer must be on the same network as the spectrum "
            "analyzers. Default IP addresses:\n\n"
            f"SA1: {DEFAULT_IP1}\n"
            f"SA2: {DEFAULT_IP2}"
        )
        
        close_btn = ctk.CTkButton(
            self,
            text="Close",
            width=120,
            height=35,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="#0066cc",
            hover_color="#0052a3",
            command=self.destroy
        )
        close_btn.pack(pady=15)
    
    def _set_focus(self):
        self.focus_force()
    
    def _add_section(self, parent, title: str, content: str):
        frame = ctk.CTkFrame(parent, fg_color="#1a1a2e", corner_radius=8)
        frame.pack(fill="x", pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            frame,
            text=title,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#00d4ff"
        )
        title_label.pack(anchor="w", padx=15, pady=(12, 5))
        
        content_label = ctk.CTkLabel(
            frame,
            text=content,
            font=ctk.CTkFont(size=12),
            text_color="#cccccc",
            justify="left",
            wraplength=450
        )
        content_label.pack(anchor="w", padx=15, pady=(0, 12))


class StatusIndicator(ctk.CTkFrame):
    """Custom status indicator with LED-style visual"""
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color="transparent")
        
        self.indicator = ctk.CTkLabel(
            self, 
            text="", 
            width=12, 
            height=12,
            corner_radius=6,
            fg_color="#666666"
        )
        self.indicator.pack(side="left", padx=(0, 8))
        
        self.status_label = ctk.CTkLabel(
            self, 
            text="Not Connected",
            font=ctk.CTkFont(size=12),
            text_color="#888888"
        )
        self.status_label.pack(side="left")
    
    def set_status(self, connected: bool, text: str):
        color = "#00ff00" if connected else "#ff4444"
        text_color = "#00ff00" if connected else "#888888"
        self.indicator.configure(fg_color=color)
        self.status_label.configure(text=text, text_color=text_color)


class SpectrumAnalyzerPanel(ctk.CTkFrame):
    """Panel for controlling a single Spectrum Analyzer with collapsible content"""
    def __init__(self, master, title: str, default_ip: str, 
                 on_connect, on_sweep, on_reset, on_continuous_toggle, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(corner_radius=10, fg_color="#1a1a2e")
        
        self.on_connect = on_connect
        self.on_sweep = on_sweep
        self.on_reset = on_reset
        self.on_continuous_toggle = on_continuous_toggle
        self.title_text = title
        self.is_collapsed = False
        self.freq_mode = "start_stop"  # "start_stop" or "center_span"
        
        header = ctk.CTkFrame(self, fg_color="#16213e", corner_radius=8)
        header.pack(fill="x", padx=10, pady=(10, 5))
        
        self.collapse_btn = ctk.CTkButton(
            header,
            text="▼",
            width=30,
            height=30,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="transparent",
            hover_color="#333366",
            text_color="#00d4ff",
            command=self._toggle_collapse
        )
        self.collapse_btn.pack(side="left", padx=(10, 0), pady=10)
        
        title_label = ctk.CTkLabel(
            header,
            text=title,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#00d4ff"
        )
        title_label.pack(side="left", padx=5, pady=10)
        
        self.status_indicator = StatusIndicator(header)
        self.status_indicator.pack(side="right", padx=15, pady=10)
        
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.pack(fill="x")
        
        connection_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        connection_frame.pack(fill="x", padx=15, pady=10)
        
        ip_label = ctk.CTkLabel(
            connection_frame, 
            text="IP Address:",
            font=ctk.CTkFont(size=13),
            width=90
        )
        ip_label.pack(side="left")
        
        self.ip_entry = ctk.CTkEntry(
            connection_frame,
            width=150,
            placeholder_text="Enter IP address",
            font=ctk.CTkFont(size=13)
        )
        self.ip_entry.insert(0, default_ip)
        self.ip_entry.pack(side="left", padx=10)
        
        self.connect_btn = ctk.CTkButton(
            connection_frame,
            text="Connect",
            width=100,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#0066cc",
            hover_color="#0052a3",
            command=self._handle_connect
        )
        self.connect_btn.pack(side="left", padx=5)
        
        settings_frame = ctk.CTkFrame(self.content_frame, fg_color="#0f0f23", corner_radius=8)
        settings_frame.pack(fill="x", padx=10, pady=5)
        
        row1 = ctk.CTkFrame(settings_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(row1, text="Points:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(0, 5))
        self.points_var = ctk.StringVar(value="3001")
        self.points_menu = ctk.CTkOptionMenu(row1, values=["101", "201", "401", "601", "801", "1001", "1501", "2001", "3001"], variable=self.points_var, width=70, font=ctk.CTkFont(size=10), fg_color="#16213e", button_color="#0066cc", button_hover_color="#0052a3")
        self.points_menu.pack(side="left", padx=0)
        
        ctk.CTkLabel(row1, text="Mode:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(15, 5))
        self.freq_mode_var = ctk.StringVar(value="Start/Stop")
        self.freq_mode_menu = ctk.CTkOptionMenu(row1, values=["Start/Stop", "Center/Span"], variable=self.freq_mode_var, width=90, font=ctk.CTkFont(size=10), fg_color="#16213e", button_color="#0066cc", button_hover_color="#0052a3", command=self._toggle_freq_mode)
        self.freq_mode_menu.pack(side="left", padx=0)
        
        ctk.CTkLabel(row1, text="BW Res:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(15, 5))
        self.bw_res_var = ctk.StringVar(value="1M")
        self.bw_res_menu = ctk.CTkOptionMenu(row1, values=["auto", "100", "1k", "10k", "100k", "1M"], variable=self.bw_res_var, width=70, font=ctk.CTkFont(size=10), fg_color="#16213e", button_color="#0066cc", button_hover_color="#0052a3")
        self.bw_res_menu.pack(side="left", padx=0)
        
        ctk.CTkLabel(row1, text="BW Vid:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(15, 5))
        self.bw_vid_var = ctk.StringVar(value="1M")
        self.bw_vid_menu = ctk.CTkOptionMenu(row1, values=["1", "10", "100", "1k", "10k", "100k", "1M", "3M"], variable=self.bw_vid_var, width=70, font=ctk.CTkFont(size=10), fg_color="#16213e", button_color="#0066cc", button_hover_color="#0052a3")
        self.bw_vid_menu.pack(side="left", padx=0)
        
        row2 = ctk.CTkFrame(settings_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        
        self.start_stop_frame = ctk.CTkFrame(row2, fg_color="transparent")
        self.start_stop_frame.pack(side="left", fill="x", expand=True)
        
        ctk.CTkLabel(self.start_stop_frame, text="Start:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(0, 3))
        self.start_entry = ctk.CTkEntry(self.start_stop_frame, width=60, font=ctk.CTkFont(size=11), justify="center")
        self.start_entry.insert(0, "0")
        self.start_entry.pack(side="left", padx=0)
        ctk.CTkLabel(self.start_stop_frame, text="MHz", font=ctk.CTkFont(size=9), text_color="#666666").pack(side="left", padx=3)
        
        ctk.CTkLabel(self.start_stop_frame, text="Stop:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(15, 3))
        self.stop_entry = ctk.CTkEntry(self.start_stop_frame, width=60, font=ctk.CTkFont(size=11), justify="center")
        self.stop_entry.insert(0, "3000")
        self.stop_entry.pack(side="left", padx=0)
        ctk.CTkLabel(self.start_stop_frame, text="MHz", font=ctk.CTkFont(size=9), text_color="#666666").pack(side="left", padx=3)
        
        self.center_span_frame = ctk.CTkFrame(row2, fg_color="transparent")
        
        ctk.CTkLabel(self.center_span_frame, text="Center:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(0, 3))
        self.center_entry = ctk.CTkEntry(self.center_span_frame, width=60, font=ctk.CTkFont(size=11), justify="center")
        self.center_entry.insert(0, "1000")
        self.center_entry.pack(side="left", padx=0)
        ctk.CTkLabel(self.center_span_frame, text="MHz", font=ctk.CTkFont(size=9), text_color="#666666").pack(side="left", padx=3)
        
        ctk.CTkLabel(self.center_span_frame, text="Span:", font=ctk.CTkFont(size=10), text_color="#888888").pack(side="left", padx=(15, 3))
        self.span_entry = ctk.CTkEntry(self.center_span_frame, width=60, font=ctk.CTkFont(size=11), justify="center")
        self.span_entry.insert(0, "2000")
        self.span_entry.pack(side="left", padx=0)
        ctk.CTkLabel(self.center_span_frame, text="MHz", font=ctk.CTkFont(size=9), text_color="#666666").pack(side="left", padx=3)
        
        buttons_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", padx=15, pady=(5, 15))
        
        self.sweep_btn = ctk.CTkButton(
            buttons_frame,
            text="Single Sweep",
            width=120,
            height=40,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#00aa44",
            hover_color="#008833",
            command=self._handle_sweep
        )
        self.sweep_btn.pack(side="left", padx=5)
        
        self.continuous_var = ctk.BooleanVar(value=False)
        self.continuous_btn = ctk.CTkButton(
            buttons_frame,
            text="Continuous",
            width=110,
            height=40,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#0066aa",
            hover_color="#005599",
            command=self._handle_continuous_toggle
        )
        self.continuous_btn.pack(side="left", padx=5)
        
        self.reset_btn = ctk.CTkButton(
            buttons_frame,
            text="Reset Peak",
            width=100,
            height=40,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#cc6600",
            hover_color="#aa5500",
            command=self._handle_reset
        )
        self.reset_btn.pack(side="left", padx=5)
        
        self.message_label = ctk.CTkLabel(
            buttons_frame,
            text="",
            font=ctk.CTkFont(size=12),
            text_color="#888888"
        )
        self.message_label.pack(side="right", padx=10)
    
    def _toggle_collapse(self):
        self.is_collapsed = not self.is_collapsed
        if self.is_collapsed:
            self.content_frame.pack_forget()
            self.collapse_btn.configure(text="▶")
        else:
            self.content_frame.pack(fill="x")
            self.collapse_btn.configure(text="▼")
    
    def _toggle_freq_mode(self, value):
        if value == "Start/Stop":
            self.freq_mode = "start_stop"
            self.center_span_frame.pack_forget()
            self.start_stop_frame.pack(side="left", fill="x", expand=True)
        else:
            self.freq_mode = "center_span"
            self.start_stop_frame.pack_forget()
            self.center_span_frame.pack(side="left", fill="x", expand=True)
    

    def _handle_connect(self):
        ip = self.ip_entry.get().strip()
        self.on_connect(ip)
    
    def _handle_sweep(self):
        self.on_sweep()
    
    def _handle_continuous_toggle(self):
        self.on_continuous_toggle()
    
    def _handle_reset(self):
        self.on_reset()
    
    def set_connected(self, connected: bool, idn: str = ""):
        if connected:
            display_idn = idn[:30] + "..." if len(idn) > 30 else idn
            self.status_indicator.set_status(True, f"Connected: {display_idn}")
            self.connect_btn.configure(text="Disconnect", fg_color="#cc4444", hover_color="#aa3333")
        else:
            self.status_indicator.set_status(False, "Not Connected")
            self.connect_btn.configure(text="Connect", fg_color="#0066cc", hover_color="#0052a3")
    
    def set_continuous_active(self, active: bool):
        if active:
            self.continuous_btn.configure(text="Stop", fg_color="#cc4444", hover_color="#aa3333")
        else:
            self.continuous_btn.configure(text="Continuous", fg_color="#0066aa", hover_color="#005599")
    
    def set_message(self, message: str, is_error: bool = False):
        color = "#ff4444" if is_error else "#00ff00"
        self.message_label.configure(text=message, text_color=color)
    
    def set_sweeping(self, sweeping: bool):
        if sweeping:
            self.sweep_btn.configure(state="disabled", text="Sweeping...")
        else:
            self.sweep_btn.configure(state="normal", text="Single Sweep")


class SpectrumPlot(ctk.CTkFrame):
    """Matplotlib-based spectrum plot with dark theme and interactive marker"""
    def __init__(self, master, title: str, on_marker_update=None, get_is_continuous=None, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(corner_radius=10, fg_color="#1a1a2e")
        
        self.title = title
        self.on_marker_update = on_marker_update
        self.get_is_continuous = get_is_continuous
        self.line_current = None
        self.line_max = None
        
        self.current_freqs = None
        self.current_vals = None
        self.current_max_vals = None
        self.previous_freq_range = None
        
        self.marker_line = None
        self.marker_dot = None
        self.marker_annotation = None
        self.marker_x = None
        self.dragging = False
        
        self.zoom_dragging = False
        self.zoom_start_x = None
        self.zoom_start_y = None
        self.zoom_rect = None
        self.zoom_history = []
        self.original_xlim = None
        self.original_ylim = None
        
        self.fig = Figure(figsize=(8, 3), dpi=100, facecolor='#1a1a2e')
        self.ax = self.fig.add_subplot(111)
        self._style_axis()
        
        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        
        self.canvas.mpl_connect('button_press_event', self._on_click)
        self.canvas.mpl_connect('motion_notify_event', self._on_motion)
        self.canvas.mpl_connect('button_release_event', self._on_release)
        
        self.fig.tight_layout(pad=2)
        
        self.marker_info_frame = ctk.CTkFrame(self, fg_color="#0f0f23", corner_radius=6)
        self.marker_info_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.marker_label = ctk.CTkLabel(
            self.marker_info_frame,
            text="Click on the spectrum to place a marker",
            font=ctk.CTkFont(size=12),
            text_color="#888888"
        )
        self.marker_label.pack(pady=8)
        
        save_buttons_frame = ctk.CTkFrame(self, fg_color="transparent")
        save_buttons_frame.pack(fill="x", padx=10, pady=(5, 10))
        
        ctk.CTkButton(
            save_buttons_frame,
            text="Save CSV",
            width=80,
            height=28,
            font=ctk.CTkFont(size=10, weight="bold"),
            fg_color="#0066aa",
            hover_color="#0052a3",
            command=self._save_csv
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            save_buttons_frame,
            text="Save Excel",
            width=80,
            height=28,
            font=ctk.CTkFont(size=10, weight="bold"),
            fg_color="#006633",
            hover_color="#005522",
            command=self._save_excel
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            save_buttons_frame,
            text="Reset Zoom",
            width=80,
            height=28,
            font=ctk.CTkFont(size=10, weight="bold"),
            fg_color="#663300",
            hover_color="#552200",
            command=self._reset_zoom
        ).pack(side="left", padx=3)
    
    def _style_axis(self):
        self.ax.set_facecolor("#0f0f23")
        self.fig.patch.set_facecolor('#1a1a2e')
        
        for spine in self.ax.spines.values():
            spine.set_color("#555588")
            spine.set_linewidth(1.5)
        
        self.ax.tick_params(axis="both", colors="#ccccdd", labelsize=11, length=6, width=1.5)
        self.ax.xaxis.label.set_color("#ddddee")
        self.ax.yaxis.label.set_color("#ddddee")
        
        self.ax.grid(True, color="#444477", linestyle="-", linewidth=0.8, alpha=0.6)
        
        self.ax.set_title(self.title, color="#00ffff", fontsize=15, fontweight="bold", pad=15)
        self.ax.set_xlabel("Frequency (MHz)", fontsize=12, fontweight="bold")
        self.ax.set_ylabel("Amplitude (dBm)", fontsize=12, fontweight="bold")
        self.ax.set_ylim(YMIN_DBM, YMAX_DBM)
    
    def _on_click(self, event):
        if event.inaxes != self.ax or self.current_freqs is None:
            return
        
        if event.button == 3:
            self.zoom_dragging = True
            self.zoom_start_x = event.xdata
            self.zoom_start_y = event.ydata
        else:
            self.dragging = True
            self._update_marker(event.xdata)
    
    def _on_motion(self, event):
        if event.inaxes != self.ax or self.current_freqs is None or event.xdata is None:
            return
        
        if self.zoom_dragging and self.zoom_start_x is not None and event.ydata is not None:
            if self.zoom_rect is not None:
                self.zoom_rect.remove()
            
            x_min = min(self.zoom_start_x, event.xdata)
            x_max = max(self.zoom_start_x, event.xdata)
            y_min = min(self.zoom_start_y, event.ydata)
            y_max = max(self.zoom_start_y, event.ydata)
            
            self.zoom_rect = Rectangle((x_min, y_min), x_max - x_min, y_max - y_min,
                                       facecolor='cyan', alpha=0.1, edgecolor='cyan', linewidth=2)
            self.ax.add_patch(self.zoom_rect)
            self.canvas.draw_idle()
        elif self.dragging:
            self._update_marker(event.xdata)
    
    def _on_release(self, event):
        if self.zoom_dragging and self.zoom_start_x is not None and event.xdata is not None:
            x_min = min(self.zoom_start_x, event.xdata)
            x_max = max(self.zoom_start_x, event.xdata)
            y_min = min(self.zoom_start_y, event.ydata)
            y_max = max(self.zoom_start_y, event.ydata)
            
            if abs(x_max - x_min) > 0.01:
                if self.original_xlim is None:
                    self.original_xlim = self.ax.get_xlim()
                if self.original_ylim is None:
                    self.original_ylim = self.ax.get_ylim()
                
                self.zoom_history.append((self.ax.get_xlim(), self.ax.get_ylim()))
                self.ax.set_xlim(x_min, x_max)
                self.ax.set_ylim(y_min, y_max)
                self.canvas.draw()
            
            if self.zoom_rect is not None:
                self.zoom_rect.remove()
                self.zoom_rect = None
            
            self.zoom_dragging = False
            self.zoom_start_x = None
            self.zoom_start_y = None
        else:
            self.dragging = False
    
    def _update_marker(self, x_pos):
        if self.current_freqs is None or self.current_vals is None:
            return
        
        if x_pos < self.current_freqs[0] or x_pos > self.current_freqs[-1]:
            return
        
        idx = np.argmin(np.abs(self.current_freqs - x_pos))
        freq = self.current_freqs[idx]
        val = self.current_vals[idx]
        max_val = self.current_max_vals[idx] if self.current_max_vals is not None else None
        
        self.marker_x = freq
        
        if self.marker_line is None:
            self.marker_line = self.ax.axvline(
                x=freq, color="#ff00ff", linewidth=1.5, linestyle="-", alpha=0.8
            )
        else:
            self.marker_line.set_xdata([freq, freq])
        
        if self.marker_dot is None:
            (self.marker_dot,) = self.ax.plot(
                freq, val, 'o', color="#ff00ff", markersize=10, zorder=10, markeredgewidth=2, markeredgecolor="#ff66ff"
            )
        else:
            self.marker_dot.set_data([freq], [val])
        
        if max_val is not None:
            marker_text = f"Freq: {freq:.3f} MHz | Current: {val:.2f} dBm | Max: {max_val:.2f} dBm"
            info_text = f"Frequency: {freq:.3f} MHz  |  Current: {val:.2f} dBm  |  Max Hold: {max_val:.2f} dBm"
        else:
            marker_text = f"Freq: {freq:.3f} MHz | Current: {val:.2f} dBm"
            info_text = f"Frequency: {freq:.3f} MHz  |  Current: {val:.2f} dBm"
        
        self.marker_label.configure(text=info_text, text_color="#00ff88")
        
        if self.on_marker_update:
            self.on_marker_update(freq, val, max_val)
        
        self.canvas.draw_idle()
    
    def update_data(self, freqs_mhz, vals, max_freqs_mhz=None, max_vals=None):
        self.current_freqs = freqs_mhz.copy()
        self.current_vals = vals.copy()
        self.current_max_vals = max_vals.copy() if max_vals is not None else None
        
        current_freq_range = (float(freqs_mhz[0]), float(freqs_mhz[-1]))
        freq_range_changed = (self.previous_freq_range is not None and 
                              (abs(self.previous_freq_range[0] - current_freq_range[0]) > 0.1 or
                               abs(self.previous_freq_range[1] - current_freq_range[1]) > 0.1))
        
        if freq_range_changed:
            self.zoom_history = []
            self.original_xlim = None
            self.original_ylim = None
        
        self.previous_freq_range = current_freq_range
        
        if self.line_current is None:
            (self.line_current,) = self.ax.plot(
                freqs_mhz, vals, 
                color="#00ff88", 
                linewidth=2.5, 
                label="Current",
                alpha=0.95
            )
        else:
            self.line_current.set_data(freqs_mhz, vals)
        
        if max_vals is not None and max_freqs_mhz is not None:
            if self.line_max is None:
                (self.line_max,) = self.ax.plot(
                    max_freqs_mhz, max_vals,
                    color="#ffff00",
                    linewidth=2.0,
                    label="Max Hold",
                    alpha=0.85
                )
            else:
                self.line_max.set_data(max_freqs_mhz, max_vals)
        
        if freq_range_changed:
            self.ax.set_xlim(current_freq_range[0], current_freq_range[1])
        else:
            self.ax.relim()
            self.ax.autoscale_view(scalex=True, scaley=False)
        
        self.ax.set_ylim(YMIN_DBM, YMAX_DBM)
        self.canvas.draw_idle()
        
        if not self.ax.get_legend():
            legend = self.ax.legend(
                loc="upper right", 
                facecolor="#1a1a2e", 
                edgecolor="#555588",
                fontsize=11,
                labelcolor="#ddddee",
                framealpha=0.95,
                fancybox=True
            )
            legend.get_frame().set_linewidth(1.5)
        
        if self.marker_x is not None:
            self._update_marker(self.marker_x)
        
        self.canvas.draw()
    
    def _save_csv(self):
        if self.get_is_continuous and self.get_is_continuous():
            messagebox.showwarning("Continuous Sweep Active", "Cannot save data while continuous sweep is running.\nStop the sweep first.")
            return
        
        if self.current_freqs is None or self.current_vals is None:
            messagebox.showwarning("No Data", "No spectrum data to save. Run a sweep first.")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"{self.title.replace(' ', '_')}_{timestamp}.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not file:
            return
        
        try:
            freqs_mhz = self.current_freqs / 1e6
            with open(file, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["Frequency (MHz)", "Amplitude (dBm)"])
                if self.current_max_vals is not None:
                    writer.writerow(["", ""])
                    writer.writerow(["Frequency (MHz)", "Peak Hold (dBm)"])
                    for freq, amp in zip(freqs_mhz, self.current_max_vals):
                        writer.writerow([f"{freq:.6f}", f"{amp:.2f}"])
                else:
                    for freq, amp in zip(freqs_mhz, self.current_vals):
                        writer.writerow([f"{freq:.6f}", f"{amp:.2f}"])
            messagebox.showinfo("Success", f"Data saved to:\n{file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save CSV:\n{str(e)}")
    
    def _save_excel(self):
        if self.get_is_continuous and self.get_is_continuous():
            messagebox.showwarning("Continuous Sweep Active", "Cannot save data while continuous sweep is running.\nStop the sweep first.")
            return
        
        if self.current_freqs is None or self.current_vals is None:
            messagebox.showwarning("No Data", "No spectrum data to save. Run a sweep first.")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing Package", "openpyxl is required for Excel export.\nInstalling...")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{self.title.replace(' ', '_')}_{timestamp}.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file:
            return
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Spectrum Data"
            
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(bold=True, color="00d4ff")
            
            ws['A1'] = "Frequency (MHz)"
            ws['B1'] = "Amplitude (dBm)"
            ws['A1'].fill = header_fill
            ws['B1'].fill = header_fill
            ws['A1'].font = header_font
            ws['B1'].font = header_font
            
            freqs_mhz = self.current_freqs / 1e6
            for idx, (freq, amp) in enumerate(zip(freqs_mhz, self.current_vals), start=2):
                ws[f'A{idx}'] = f"{freq:.6f}"
                ws[f'B{idx}'] = f"{amp:.2f}"
            
            if self.current_max_vals is not None:
                row = len(freqs_mhz) + 4
                ws[f'A{row}'] = "Frequency (MHz)"
                ws[f'B{row}'] = "Peak Hold (dBm)"
                ws[f'A{row}'].fill = header_fill
                ws[f'B{row}'].fill = header_fill
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].font = header_font
                
                for idx, (freq, amp) in enumerate(zip(freqs_mhz, self.current_max_vals), start=row+1):
                    ws[f'A{idx}'] = f"{freq:.6f}"
                    ws[f'B{idx}'] = f"{amp:.2f}"
            
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            
            wb.save(file)
            messagebox.showinfo("Success", f"Data saved to:\n{file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{str(e)}")
    
    def _reset_zoom(self):
        if self.original_xlim is not None:
            self.ax.set_xlim(self.original_xlim)
        else:
            if self.current_freqs is not None:
                self.ax.set_xlim(self.current_freqs[0], self.current_freqs[-1])
        self.ax.set_ylim(YMIN_DBM, YMAX_DBM)
        self.zoom_history = []
        self.original_xlim = None
        self.original_ylim = None
        self.canvas.draw()
    
    def clear(self):
        self.ax.clear()
        self._style_axis()
        self.line_current = None
        self.line_max = None
        self.marker_line = None
        self.marker_dot = None
        self.marker_annotation = None
        self.marker_x = None
        self.current_freqs = None
        self.current_vals = None
        self.current_max_vals = None
        self.previous_freq_range = None
        self.zoom_history = []
        self.original_xlim = None
        self.original_ylim = None
        self.marker_label.configure(
            text="Click on the spectrum to place a marker",
            text_color="#888888"
        )
        self.canvas.draw()


class DualSAViewer(ctk.CTk):
    """Main application window for Dual Spectrum Analyzer Viewer"""
    
    def __init__(self):
        super().__init__()
        
        self.title("Dual Spectrum Analyzer Viewer - DSA1030")
        self.geometry("1200x900")
        self.minsize(1000, 700)
        
        self.rm = None
        self.inst1 = None
        self.inst2 = None
        
        self.max_vals1 = None
        self.max_freqs1 = None
        self.max_vals2 = None
        self.max_freqs2 = None
        
        self.continuous1 = False
        self.continuous2 = False
        self.sweep_thread1 = None
        self.sweep_thread2 = None
        
        self.fullscreen_mode = None
        
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        self._build_ui()
    
    def _build_ui(self):
        self.configure(fg_color="#0a0a1a")
        
        header_frame = ctk.CTkFrame(self, fg_color="#16213e", corner_radius=0, height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        title_label = ctk.CTkLabel(
            header_frame,
            text="Dual Spectrum Analyzer Viewer",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="#00d4ff"
        )
        title_label.pack(side="left", padx=30, pady=15)
        
        subtitle = ctk.CTkLabel(
            header_frame,
            text="DSA1030 Controller",
            font=ctk.CTkFont(size=14),
            text_color="#666688"
        )
        subtitle.pack(side="left", padx=10, pady=15)
        
        view_buttons_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        view_buttons_frame.pack(side="right", padx=10, pady=10)
        
        self.fs1_btn = ctk.CTkButton(
            view_buttons_frame,
            text="SA1 ⛶",
            width=70,
            height=32,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#224444",
            hover_color="#335555",
            command=lambda: self._toggle_fullscreen(1)
        )
        self.fs1_btn.pack(side="left", padx=4)
        
        self.fs2_btn = ctk.CTkButton(
            view_buttons_frame,
            text="SA2 ⛶",
            width=70,
            height=32,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#443344",
            hover_color="#554455",
            command=lambda: self._toggle_fullscreen(2)
        )
        self.fs2_btn.pack(side="left", padx=4)
        
        self.split_btn = ctk.CTkButton(
            view_buttons_frame,
            text="Split",
            width=60,
            height=32,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#333366",
            hover_color="#444488",
            command=lambda: self._toggle_fullscreen(None)
        )
        self.split_btn.pack(side="left", padx=4)
        
        help_btn = ctk.CTkButton(
            view_buttons_frame,
            text="?",
            width=40,
            height=40,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color="#333366",
            hover_color="#444488",
            corner_radius=20,
            command=self._show_help
        )
        help_btn.pack(side="left", padx=5)
        
        doc_btn = ctk.CTkButton(
            view_buttons_frame,
            text="📖",
            width=40,
            height=40,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color="#444433",
            hover_color="#555544",
            corner_radius=20,
            command=self._open_documentation
        )
        doc_btn.pack(side="left", padx=2)
        
        reset_net_btn = ctk.CTkButton(
            view_buttons_frame,
            text="🔄",
            width=40,
            height=40,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color="#663333",
            hover_color="#774444",
            corner_radius=20,
            command=self._reset_network
        )
        reset_net_btn.pack(side="left", padx=2)
        
        self.control_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.control_frame.pack(fill="x", padx=20, pady=(10, 5))
        
        main_container = ctk.CTkScrollableFrame(self, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        left_control = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        left_control.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        right_control = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        right_control.pack(side="right", fill="x", expand=True, padx=(10, 0))
        
        self.sa1_panel = SpectrumAnalyzerPanel(
            left_control,
            title="Spectrum Analyzer 1",
            default_ip=DEFAULT_IP1,
            on_connect=self.connect1,
            on_sweep=self.sweep1,
            on_reset=self.reset_peak1,
            on_continuous_toggle=self.toggle_continuous1
        )
        self.sa1_panel.pack(fill="x")
        
        self.sa2_panel = SpectrumAnalyzerPanel(
            right_control,
            title="Spectrum Analyzer 2",
            default_ip=DEFAULT_IP2,
            on_connect=self.connect2,
            on_sweep=self.sweep2,
            on_reset=self.reset_peak2,
            on_continuous_toggle=self.toggle_continuous2
        )
        self.sa2_panel.pack(fill="x")
        
        self.plots_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        self.plots_frame.pack(fill="both", expand=True)
        
        self.plot1_frame = ctk.CTkFrame(self.plots_frame, fg_color="transparent")
        self.plot1_frame.pack(fill="both", expand=True, pady=(0, 3))
        
        self.plot1 = SpectrumPlot(self.plot1_frame, title="SA1 Spectrum", on_marker_update=self._update_sa1_marker, get_is_continuous=lambda: self.continuous1)
        self.plot1.pack(fill="both", expand=True)
        
        divider = ctk.CTkFrame(self.plots_frame, fg_color="#333366", height=6)
        divider.pack(fill="x", padx=0, pady=1)
        divider.bind("<Enter>", lambda e: divider.configure(fg_color="#4444aa"))
        divider.bind("<Leave>", lambda e: divider.configure(fg_color="#333366") if not self.divider_dragging else None)
        divider.bind("<Button-1>", self._on_divider_press)
        divider.bind("<B1-Motion>", self._on_divider_drag)
        divider.bind("<ButtonRelease-1>", self._on_divider_release)
        self.divider = divider
        self.divider_dragging = False
        self.divider_start_y = 0
        
        self.plot2_frame = ctk.CTkFrame(self.plots_frame, fg_color="transparent")
        self.plot2_frame.pack(fill="both", expand=True, pady=(3, 0))
        
        self.plot2 = SpectrumPlot(self.plot2_frame, title="SA2 Spectrum", on_marker_update=self._update_sa2_marker, get_is_continuous=lambda: self.continuous2)
        self.plot2.pack(fill="both", expand=True)
        
        footer = ctk.CTkFrame(self, fg_color="#16213e", corner_radius=0, height=40)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        
        self.footer_text = ctk.CTkLabel(
            footer,
            text="Ready",
            font=ctk.CTkFont(size=11),
            text_color="#666688"
        )
        self.footer_text.pack(side="left", padx=20, pady=8)
        
        marker_info_container = ctk.CTkFrame(footer, fg_color="transparent")
        marker_info_container.pack(side="right", padx=20, pady=5)
        
        self.sa2_marker_label = ctk.CTkLabel(
            marker_info_container,
            text="SA2: --",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#ff66aa"
        )
        self.sa2_marker_label.pack(side="right", padx=(15, 0))
        
        self.sa1_marker_label = ctk.CTkLabel(
            marker_info_container,
            text="SA1: --",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#00ff88"
        )
        self.sa1_marker_label.pack(side="right", padx=0)
    
    def _toggle_fullscreen(self, mode):
        if self.fullscreen_mode == mode:
            return
        
        self.fullscreen_mode = mode
        
        if mode is None:
            self.plot1_frame.configure(height=-1)
            self.plot2_frame.configure(height=-1)
            self.plot1_frame.pack_propagate(True)
            self.plot2_frame.pack_propagate(True)
            self.divider.pack(fill="x", padx=0, pady=2)
            self.split_btn.configure(fg_color="#444488", hover_color="#555599")
            self.fs1_btn.configure(fg_color="#224444", hover_color="#335555")
            self.fs2_btn.configure(fg_color="#443344", hover_color="#554455")
        elif mode == 1:
            self.plot1_frame.pack_propagate(False)
            self.plot2_frame.pack_propagate(False)
            self.plot1_frame.configure(height=600)
            self.plot2_frame.configure(height=100)
            self.divider.pack_forget()
            self.fs1_btn.configure(fg_color="#00ff88", hover_color="#00dd77")
            self.fs2_btn.configure(fg_color="#443344", hover_color="#554455")
            self.split_btn.configure(fg_color="#333366", hover_color="#444488")
        elif mode == 2:
            self.plot1_frame.pack_propagate(False)
            self.plot2_frame.pack_propagate(False)
            self.plot1_frame.configure(height=100)
            self.plot2_frame.configure(height=600)
            self.divider.pack_forget()
            self.fs2_btn.configure(fg_color="#ff66aa", hover_color="#ff55aa")
            self.fs1_btn.configure(fg_color="#224444", hover_color="#335555")
            self.split_btn.configure(fg_color="#333366", hover_color="#444488")
    
    def _on_divider_press(self, event):
        self.divider_dragging = True
        self.divider_start_y = event.y_root
        self.divider.configure(fg_color="#6666ff")
    
    def _on_divider_drag(self, event):
        if not self.divider_dragging or self.fullscreen_mode is not None:
            return
        
        delta = event.y_root - self.divider_start_y
        if abs(delta) < 2:
            return
        
        self.divider_start_y = event.y_root
        
        try:
            plot1_height = self.plot1_frame.winfo_height()
            plot2_height = self.plot2_frame.winfo_height()
            new_plot1_height = max(150, plot1_height + delta)
            new_plot2_height = max(150, plot2_height - delta)
            
            if new_plot1_height > 150 and new_plot2_height > 150:
                self.plot1_frame.configure(height=new_plot1_height)
                self.plot2_frame.configure(height=new_plot2_height)
                self.plot1_frame.pack_propagate(False)
                self.plot2_frame.pack_propagate(False)
        except Exception as e:
            pass
    
    def _on_divider_release(self, event):
        self.divider_dragging = False
        self.divider.configure(fg_color="#333366")
    
    def _show_help(self):
        HelpDialog(self)
    
    def _open_documentation(self):
        pdf_path = os.path.join(os.path.dirname(__file__), "attached_assets", "DSA1000 ProgrammingGuide_1764286444318.pdf")
        if os.path.exists(pdf_path):
            webbrowser.open('file://' + os.path.realpath(pdf_path))
        else:
            messagebox.showwarning("Documentation", "PDF documentation file not found.")
    
    def _reset_network(self):
        print("Resetting VISA resource manager...")
        self.continuous1 = False
        self.continuous2 = False
        self.sa1_panel.set_continuous_active(False)
        self.sa2_panel.set_continuous_active(False)
        
        try:
            if self.inst1 is not None:
                self.inst1.close()
            self.inst1 = None
        except:
            pass
        
        try:
            if self.inst2 is not None:
                self.inst2.close()
            self.inst2 = None
        except:
            pass
        
        try:
            if self.rm is not None:
                self.rm.close()
            self.rm = None
        except:
            pass
        
        self.sa1_panel.set_connected(False)
        self.sa2_panel.set_connected(False)
        self.sa1_panel.set_message("Network reset - ready to reconnect")
        self.sa2_panel.set_message("Network reset - ready to reconnect")
        self.plot1.clear()
        self.plot2.clear()
        self.max_vals1 = None
        self.max_freqs1 = None
        self.max_vals2 = None
        self.max_freqs2 = None
        print("Network reset complete")
    
    def _update_sa1_marker(self, freq, val, max_val):
        if max_val is not None:
            text = f"SA1: {freq:.2f} MHz | {val:.1f} dBm (Max: {max_val:.1f})"
        else:
            text = f"SA1: {freq:.2f} MHz | {val:.1f} dBm"
        self.sa1_marker_label.configure(text=text)
    
    def _update_sa2_marker(self, freq, val, max_val):
        if max_val is not None:
            text = f"SA2: {freq:.2f} MHz | {val:.1f} dBm (Max: {max_val:.1f})"
        else:
            text = f"SA2: {freq:.2f} MHz | {val:.1f} dBm"
        self.sa2_marker_label.configure(text=text)
    

    def _on_closing(self):
        self.continuous1 = False
        self.continuous2 = False
        
        self.footer_text.configure(text="Closing connections...")
        self.update()
        
        if self.inst1 is not None:
            try:
                self.inst1.close()
            except:
                pass
            self.inst1 = None
        
        if self.inst2 is not None:
            try:
                self.inst2.close()
            except:
                pass
            self.inst2 = None
        
        if self.rm is not None:
            try:
                self.rm.close()
            except:
                pass
            self.rm = None
        
        self.destroy()
    
    def _get_rm(self):
        """
        Create and return a pyvisa ResourceManager.
    
        Try the default ResourceManager() first (this will use NI-VISA on Windows
        if it's installed). If that fails, fall back to the pure-Python backend
        provided by pyvisa-py via ResourceManager('@py').
    
        If neither backend can be created, raise a RuntimeError so the calling
        code can handle it (and so the error is obvious during startup).
        """
        if self.rm is not None:
            return self.rm
    
        # Try the default (system) VISA backend first
        try:
            self.rm = pyvisa.ResourceManager()
            # Try a lightweight check (list_resources may succeed or raise)
            try:
                _ = self.rm.list_resources()
            except Exception:
                # listing may fail depending on backend configuration; ignore here
                pass
            print("Using default VISA ResourceManager()")
            return self.rm
        except Exception as e_default:
            print(f"default ResourceManager() failed: {e_default}. Trying pyvisa-py ('@py')...")
    
        # Fallback to pyvisa-py
        try:
            self.rm = pyvisa.ResourceManager("@py")
            try:
                _ = self.rm.list_resources()
            except Exception:
                pass
            print("Using pyvisa-py ResourceManager('@py')")
            return self.rm
        except Exception as e_py:
            print(f"pyvisa-py ResourceManager('@py') failed: {e_py}")
            self.rm = None
            raise RuntimeError("No usable VISA backend found. Install NI-VISA or pyvisa-py and try again.") from e_py
        
    def _check_connection1(self):
        if self.inst1 is None:
            return False
        try:
            self.inst1.write("*OPC?")
            time.sleep(0.1)
            response = self.inst1.read()
            return "1" in response
        except Exception as e:
            print(f"SA1 connection check failed: {e}")
            return False
    
    def _check_connection2(self):
        if self.inst2 is None:
            return False
        try:
            self.inst2.write("*OPC?")
            time.sleep(0.1)
            response = self.inst2.read()
            return "1" in response
        except Exception as e:
            print(f"SA2 connection check failed: {e}")
            return False
    
    def connect1(self, ip: str):
        if not ip:
            messagebox.showerror("SA1 Error", "Please enter a valid IP address")
            return
        
        if self.inst1 is not None:
            self.continuous1 = False
            self.sa1_panel.set_continuous_active(False)
            try:
                self.inst1.close()
            except:
                pass
            self.inst1 = None
            self.sa1_panel.set_connected(False)
            self.sa1_panel.set_message("Disconnected")
            self.max_vals1 = None
            self.max_freqs1 = None
            self.plot1.clear()
            return
        
        try:
            print(f"SA1: Attempting to connect to {ip}...")
            rm = self._get_rm()
            print(f"SA1: Resource manager created, opening resource...")
            self.inst1 = rm.open_resource(f"TCPIP::{ip}::INSTR")
            print(f"SA1: Resource opened, setting timeout to 120000ms...")
            self.inst1.timeout = 120000
            print(f"SA1: Waiting for device to stabilize...")
            time.sleep(2)
            print(f"SA1: Querying device ID...")
            self.inst1.write("*IDN?")
            time.sleep(0.5)
            idn = self.inst1.read().strip()
            print(f"SA1: Device ID: {idn}")
            self.sa1_panel.set_connected(True, idn)
            self.sa1_panel.set_message("Connected successfully")
        except Exception as e:
            error_str = f"{type(e).__name__}: {str(e)}"
            print(f"SA1 Connection Error: {error_str}")
            self.inst1 = None
            self.sa1_panel.set_connected(False)
            self.sa1_panel.set_message(f"Connection failed", is_error=True)
            messagebox.showerror("SA1 Error", error_str)
    
    def connect2(self, ip: str):
        if not ip:
            messagebox.showerror("SA2 Error", "Please enter a valid IP address")
            return
        
        if self.inst2 is not None:
            self.continuous2 = False
            self.sa2_panel.set_continuous_active(False)
            try:
                self.inst2.close()
            except:
                pass
            self.inst2 = None
            self.sa2_panel.set_connected(False)
            self.sa2_panel.set_message("Disconnected")
            self.max_vals2 = None
            self.max_freqs2 = None
            self.plot2.clear()
            return
        
        try:
            print(f"SA2: Attempting to connect to {ip}...")
            rm = self._get_rm()
            print(f"SA2: Resource manager created, opening resource...")
            self.inst2 = rm.open_resource(f"TCPIP::{ip}::INSTR")
            print(f"SA2: Resource opened, setting timeout to 120000ms...")
            self.inst2.timeout = 120000
            print(f"SA2: Waiting for device to stabilize...")
            time.sleep(2)
            print(f"SA2: Querying device ID...")
            self.inst2.write("*IDN?")
            time.sleep(0.5)
            idn = self.inst2.read().strip()
            print(f"SA2: Device ID: {idn}")
            self.sa2_panel.set_connected(True, idn)
            self.sa2_panel.set_message("Connected successfully")
        except Exception as e:
            error_str = f"{type(e).__name__}: {str(e)}"
            print(f"SA2 Connection Error: {error_str}")
            self.inst2 = None
            self.sa2_panel.set_connected(False)
            self.sa2_panel.set_message(f"Connection failed", is_error=True)
            messagebox.showerror("SA2 Error", error_str)
    
    def _configure_sa(self, inst, start_mhz, stop_mhz, points=3001, bw_res="10k", bw_vid="10k"):
        start_hz = start_mhz * 1e6
        stop_hz = stop_mhz * 1e6
        center_hz = (start_hz + stop_hz) / 2
        span_hz = stop_hz - start_hz
        
        try:
            if bw_res == "auto":
                inst.write(":BAND:RES:AUTO ON")
            else:
                bw_res_hz = float(bw_res.replace('k', 'e3').replace('M', 'e6'))
                inst.write(f":BAND:RES {bw_res_hz}")
            
            bw_vid_hz = float(bw_vid.replace('k', 'e3').replace('M', 'e6'))
            
            inst.write(f":FREQ:STAR {start_hz}")
            time.sleep(0.1)
            inst.write(f":FREQ:STOP {stop_hz}")
            time.sleep(0.1)
            inst.write(f":FREQ:CENT {center_hz}")
            time.sleep(0.1)
            inst.write(f":FREQ:SPAN {span_hz}")
            time.sleep(0.1)
            inst.write(f":SWE:POIN {points}")
            time.sleep(0.1)
            inst.write(f":BAND:VID {bw_vid_hz}")
            time.sleep(0.1)
            inst.write(":DET POS")
            time.sleep(0.1)
            inst.write(":FORM:TRAC:DATA ASCii")
            time.sleep(0.1)
        except Exception as e:
            print(f"Error configuring device: {e}")
    
    def _single_sweep(self, inst):
        inst.write(":INIT:CONT OFF")
        time.sleep(0.1)
        inst.write(":INIT")
        try:
            inst.write(":SWE:TIME?")
            time.sleep(0.15)
            st = float(inst.read())
        except Exception:
            st = 1.5
        time.sleep(max(1.5, st * 1.1))
    
    def _read_trace_ascii_block(self, inst):
        inst.write(":SWE:POIN?")
        time.sleep(0.2)
        points = int(inst.read().strip())
        
        inst.write(":FREQ:STAR?")
        time.sleep(0.2)
        f_start = float(inst.read().strip())
        
        inst.write(":FREQ:STOP?")
        time.sleep(0.2)
        f_stop = float(inst.read().strip())
        
        inst.write(":FORM:TRAC:DATA ASCii")
        time.sleep(0.2)
        inst.write(":TRAC:DATA? TRACE1")
        time.sleep(0.3)
        raw = inst.read().strip()
        
        if raw.startswith("#"):
            n_digits = int(raw[1])
            len_str = raw[2:2 + n_digits]
            data_len = int(len_str)
            header_len = 2 + n_digits
            data_str = raw[header_len:header_len + data_len]
        else:
            data_str = raw
        
        parts = [p for p in data_str.split(",") if p.strip()]
        vals = np.array([float(p) for p in parts], dtype=float)
        
        if vals.size != points:
            print(f"Warning: expected {points} points, got {vals.size}")
        
        freqs = np.linspace(f_start, f_stop, num=vals.size)
        return freqs, vals
    
    def reset_peak1(self):
        self.max_vals1 = None
        self.max_freqs1 = None
        self.sa1_panel.set_message("Peak hold reset")
    
    def sweep1(self, center: str = None, span: str = None, points: str = None):
        if self.inst1 is None:
            messagebox.showerror("SA1 Error", "SA1 not connected")
            return
        
        def do_sweep():
            try:
                start_val, stop_val = self._get_freq_range_for_panel(self.sa1_panel)
                pts = int(self.sa1_panel.points_var.get())
                bw_res = self.sa1_panel.bw_res_var.get()
                bw_vid = self.sa1_panel.bw_vid_var.get()
                
                self.after(0, lambda: self.sa1_panel.set_sweeping(True))
                
                self._configure_sa(self.inst1, start_val, stop_val, pts, bw_res, bw_vid)
                self._single_sweep(self.inst1)
                freqs, vals = self._read_trace_ascii_block(self.inst1)
                
                self.after(0, lambda: self._update_plot1(freqs, vals))
                self.after(0, lambda: self.sa1_panel.set_message("Sweep completed"))
                self.after(0, lambda: self.sa1_panel.set_sweeping(False))
            except Exception as e:
                error_msg = str(e)
                self.after(0, lambda: self.sa1_panel.set_message("Sweep error", is_error=True))
                self.after(0, lambda: self.sa1_panel.set_sweeping(False))
                self.after(0, lambda msg=error_msg: messagebox.showerror("SA1 Error", msg))
        
        thread = threading.Thread(target=do_sweep, daemon=True)
        thread.start()
    
    def toggle_continuous1(self):
        if self.inst1 is None:
            messagebox.showerror("SA1 Error", "SA1 not connected")
            return
        
        self.continuous1 = not self.continuous1
        self.sa1_panel.set_continuous_active(self.continuous1)
        
        if self.continuous1:
            self._start_continuous1()
    
    def _start_continuous1(self):
        if not self.continuous1 or self.inst1 is None:
            return
        
        def do_continuous():
            sweep_count = 0
            try:
                start_val, stop_val = self._get_freq_range_for_panel(self.sa1_panel)
                pts = int(self.sa1_panel.points_var.get())
                bw_res = self.sa1_panel.bw_res_var.get()
                bw_vid = self.sa1_panel.bw_vid_var.get()
                
                self._configure_sa(self.inst1, start_val, stop_val, pts, bw_res, bw_vid)
                self.inst1.write(":INIT:CONT OFF")
                time.sleep(0.3)
                
                original_timeout = self.inst1.timeout
                self.inst1.timeout = 30000
                
                while self.continuous1 and self.inst1 is not None:
                    try:
                        self.inst1.write(":INIT")
                        self.inst1.write(":SWE:TIME?")
                        time.sleep(0.2)
                        try:
                            sweep_time = float(self.inst1.read())
                        except:
                            sweep_time = 1.5
                        time.sleep(max(1.5, sweep_time * 1.2))
                        freqs, vals = self._read_trace_ascii_block(self.inst1)
                        self.after(0, lambda f=freqs, v=vals: self._update_plot1(f, v))
                        sweep_count += 1
                    except Exception as sweep_err:
                        print(f"Sweep error (sweep #{sweep_count}): {sweep_err}")
                        time.sleep(0.5)
                        break
                
                self.inst1.timeout = original_timeout
            except Exception as e:
                print(f"Continuous sweep setup error: {e}")
                self.after(0, lambda: self.sa1_panel.set_message("Sweep error", is_error=True))
            finally:
                try:
                    self.inst1.write(":INIT:CONT OFF")
                except:
                    pass
                self.continuous1 = False
                self.after(0, lambda: self.sa1_panel.set_continuous_active(False))
        
        thread = threading.Thread(target=do_continuous, daemon=True)
        thread.start()
    
    def _update_plot1(self, freqs, vals):
        freqs_mhz = freqs / 1e6
        
        if self.max_vals1 is None or self.max_freqs1 is None or len(self.max_vals1) != len(vals):
            self.max_freqs1 = freqs.copy()
            self.max_vals1 = vals.copy()
        else:
            self.max_vals1 = np.maximum(self.max_vals1, vals)
        
        max_freqs_mhz = self.max_freqs1 / 1e6
        self.plot1.update_data(freqs_mhz, vals, max_freqs_mhz, self.max_vals1)
    
    def _get_freq_range_for_panel(self, panel):
        """Get start and stop MHz values based on current frequency mode"""
        if panel.freq_mode == "center_span":
            center = float(panel.center_entry.get().replace(",", "."))
            span = float(panel.span_entry.get().replace(",", "."))
            start = center - span / 2
            stop = center + span / 2
            return start, stop
        else:
            start = float(panel.start_entry.get().replace(",", "."))
            stop = float(panel.stop_entry.get().replace(",", "."))
            return start, stop
    
    def reset_peak1(self):
        self.max_vals1 = None
        self.max_freqs1 = None
        if self.plot1 is not None and self.plot1.line_max is not None:
            try:
                self.plot1.line_max.remove()
            except:
                pass
            self.plot1.line_max = None
            self.plot1.canvas.draw_idle()
        self.sa1_panel.set_message("Peak hold reset")
    
    def reset_peak2(self):
        self.max_vals2 = None
        self.max_freqs2 = None
        if self.plot2 is not None and self.plot2.line_max is not None:
            try:
                self.plot2.line_max.remove()
            except:
                pass
            self.plot2.line_max = None
            self.plot2.canvas.draw_idle()
        self.sa2_panel.set_message("Peak hold reset")
    
    def sweep2(self, center: str = None, span: str = None, points: str = None):
        if self.inst2 is None:
            messagebox.showerror("SA2 Error", "SA2 not connected")
            return
        
        def do_sweep():
            try:
                start_val, stop_val = self._get_freq_range_for_panel(self.sa2_panel)
                pts = int(self.sa2_panel.points_var.get())
                bw_res = self.sa2_panel.bw_res_var.get()
                bw_vid = self.sa2_panel.bw_vid_var.get()
                
                self.after(0, lambda: self.sa2_panel.set_sweeping(True))
                
                self._configure_sa(self.inst2, start_val, stop_val, pts, bw_res, bw_vid)
                self._single_sweep(self.inst2)
                freqs, vals = self._read_trace_ascii_block(self.inst2)
                
                self.after(0, lambda: self._update_plot2(freqs, vals))
                self.after(0, lambda: self.sa2_panel.set_sweeping(False))
                self.after(0, lambda: self.sa2_panel.set_message("Sweep completed"))
            except Exception as e:
                error_msg = str(e)
                self.after(0, lambda: self.sa2_panel.set_sweeping(False))
                self.after(0, lambda: self.sa2_panel.set_message("Sweep error", is_error=True))
                self.after(100, lambda msg=error_msg: messagebox.showerror("SA2 Error", msg))
        
        thread = threading.Thread(target=do_sweep, daemon=True)
        thread.start()
    
    def toggle_continuous2(self):
        if self.inst2 is None:
            messagebox.showerror("SA2 Error", "SA2 not connected")
            return
        
        self.continuous2 = not self.continuous2
        self.sa2_panel.set_continuous_active(self.continuous2)
        
        if self.continuous2:
            self._start_continuous2()
    
    def _start_continuous2(self):
        if not self.continuous2 or self.inst2 is None:
            return
        
        def do_continuous():
            sweep_count = 0
            try:
                start_val, stop_val = self._get_freq_range_for_panel(self.sa2_panel)
                pts = int(self.sa2_panel.points_var.get())
                bw_res = self.sa2_panel.bw_res_var.get()
                bw_vid = self.sa2_panel.bw_vid_var.get()
                
                self._configure_sa(self.inst2, start_val, stop_val, pts, bw_res, bw_vid)
                self.inst2.write(":INIT:CONT OFF")
                time.sleep(0.3)
                
                original_timeout = self.inst2.timeout
                self.inst2.timeout = 30000
                
                while self.continuous2 and self.inst2 is not None:
                    try:
                        self.inst2.write(":INIT")
                        self.inst2.write(":SWE:TIME?")
                        time.sleep(0.2)
                        try:
                            sweep_time = float(self.inst2.read())
                        except:
                            sweep_time = 1.5
                        time.sleep(max(1.5, sweep_time * 1.2))
                        freqs, vals = self._read_trace_ascii_block(self.inst2)
                        self.after(0, lambda f=freqs, v=vals: self._update_plot2(f, v))
                        sweep_count += 1
                    except Exception as sweep_err:
                        print(f"Sweep error (sweep #{sweep_count}): {sweep_err}")
                        time.sleep(0.5)
                        break
                
                self.inst2.timeout = original_timeout
            except Exception as e:
                print(f"Continuous sweep setup error: {e}")
                self.after(0, lambda: self.sa2_panel.set_message("Sweep error", is_error=True))
            finally:
                try:
                    self.inst2.write(":INIT:CONT OFF")
                except:
                    pass
                self.continuous2 = False
                self.after(0, lambda: self.sa2_panel.set_continuous_active(False))
        
        thread = threading.Thread(target=do_continuous, daemon=True)
        thread.start()
    
    def _update_plot2(self, freqs, vals):
        freqs_mhz = freqs / 1e6
        
        if self.max_vals2 is None or self.max_freqs2 is None or len(self.max_vals2) != len(vals):
            self.max_freqs2 = freqs.copy()
            self.max_vals2 = vals.copy()
        else:
            self.max_vals2 = np.maximum(self.max_vals2, vals)
        
        max_freqs_mhz = self.max_freqs2 / 1e6
        self.plot2.update_data(freqs_mhz, vals, max_freqs_mhz, self.max_vals2)


def main():
    app = DualSAViewer()
    app.mainloop()


if __name__ == "__main__":
    main()
